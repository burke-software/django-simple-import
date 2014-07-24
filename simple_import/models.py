from django.contrib.contenttypes.models import ContentType
from django.contrib.contenttypes import generic
from django.db import models
from django.conf import settings
from django.db import transaction
from django.utils.encoding import smart_text
import datetime

AUTH_USER_MODEL = getattr(settings, 'AUTH_USER_MODEL', 'auth.User')

import sys
if sys.version_info >= (3,0):
    unicode = str

class ImportSetting(models.Model):
    """ Save some settings per user per content type """
    user = models.ForeignKey(AUTH_USER_MODEL)
    content_type = models.ForeignKey(ContentType)
    
    class Meta():
        unique_together = ('user', 'content_type',)


class ColumnMatch(models.Model):
    """ Match column names from the user uploaded file to the database """
    column_name = models.CharField(max_length=200)
    field_name = models.CharField(max_length=255, blank=True)
    import_setting = models.ForeignKey(ImportSetting)
    default_value = models.CharField(max_length=2000, blank=True)
    null_on_empty = models.BooleanField(default=False, help_text="If cell is blank, clear out the field setting it to blank.")
    header_position = models.IntegerField(help_text="Annoying way to order the columns to match the header rows")
    
    class Meta:
        unique_together = ('column_name', 'import_setting')
    
    def __unicode__(self):
        return unicode('{0} {1}'.format(self.column_name, self.field_name))

    def guess_field(self):
        """ Guess the match based on field names
        First look for an exact field name match
        then search defined alternative names
        then normalize the field name and check for match
        """
        model = self.import_setting.content_type.model_class()
        field_names = model._meta.get_all_field_names()
        if self.column_name in field_names:
            self.field_name = self.column_name
            return
        #TODO user defined alt names
        normalized_field_name = self.column_name.lower().replace(' ', '_')
        if normalized_field_name in field_names:
            self.field_name = self.column_name
        # Try verbose name
        for field_name in field_names:
            field = model._meta.get_field_by_name(field_name)[0]
            if hasattr(field, 'verbose_name'):
                if field.verbose_name.lower().replace(' ', '_') == normalized_field_name:
                    self.field_name = field_name


class ImportLog(models.Model):
    """ A log of all import attempts """
    name = models.CharField(max_length=255)
    user = models.ForeignKey(AUTH_USER_MODEL, editable=False, related_name="simple_import_log")
    date = models.DateTimeField(auto_now_add=True, verbose_name="Date Created")
    import_file = models.FileField(upload_to="import_file")
    error_file = models.FileField(upload_to="error_file", blank=True)
    import_setting = models.ForeignKey(ImportSetting, editable=False)
    import_type_choices = (
        ("N", "Create New Records"),
        ("U", "Create and Update Records"),
        ("O", "Only Update Records"),
    )
    import_type = models.CharField(max_length=1, choices=import_type_choices)
    update_key = models.CharField(max_length=200, blank=True)
    
    def __unicode__(self):
        return unicode(self.name)
    
    def clean(self):
        from django.core.exceptions import ValidationError
        filename = str(self.import_file).lower()
        if not filename[-3:] in ('xls', 'ods', 'csv', 'lsx'):
            raise ValidationError('Invalid file type. Must be xls, xlsx, ods, or csv.')
    
    @transaction.commit_on_success
    def undo(self):
        if self.import_type != "N":
            raise Exception("Cannot undo this type of import!")
        for obj in self.importedobject_set.all():
            if obj.content_object:
                obj.content_object.delete()
            obj.delete()

    @staticmethod
    def is_empty(value):
        """ Check `value` for emptiness by first comparing with None and then
        by coercing to string, trimming, and testing for zero length """
        return value is None or not len(smart_text(value).strip())
    
    def get_matches(self):
        """ Get each matching header row to database match
        Returns a ColumnMatch queryset"""
        header_row = self.get_import_file_as_list(only_header=True)
        match_ids = []
        
        for i, cell in enumerate(header_row):
            if self.is_empty(cell): # Sometimes we get blank headers, ignore them.
                continue
            
            try:
                match = ColumnMatch.objects.get(
                    import_setting = self.import_setting,
                    column_name = cell,
                )
            except ColumnMatch.DoesNotExist:
                match = ColumnMatch(
                    import_setting = self.import_setting,
                    column_name = cell,
                )
                match.guess_field()
            
            match.header_position = i
            match.save()
            
            match_ids += [match.id]
        
        return ColumnMatch.objects.filter(pk__in=match_ids).order_by('header_position')

    def get_import_file_as_list(self, only_header=False):
        file_ext = str(self.import_file).lower()[-3:]
        data = []
        
        self.import_file.seek(0)
        
        if file_ext == "xls":
            import xlrd
            import os
            
            wb = xlrd.open_workbook(file_contents=self.import_file.read())
            sh1 = wb.sheet_by_index(0)
            for rownum in range(sh1.nrows):
                row_values = []
                for cell in sh1.row(rownum):
                    # xlrd is too dumb to just check for dates. So we have to ourselves
                    if cell.ctype == 3: # 3 is date - http://www.lexicon.net/sjmachin/xlrd.html#xlrd.Cell-class
                        row_values += [datetime.datetime(*xlrd.xldate_as_tuple(cell.value, wb.datemode))]
                    else:
                        row_values += [cell.value]
                data += [row_values]
                if only_header:
                    break
        elif file_ext == "csv":
            import csv
            reader = csv.reader(self.import_file)
            for row in reader:
                data += [row]
                if only_header:
                    break
        elif file_ext == "lsx":
            from openpyxl.reader.excel import load_workbook
            # load_workbook actually accepts a file-like object for the filename param
            wb = load_workbook(filename=self.import_file, use_iterators = True)
            sheet = wb.get_active_sheet()
            for row in sheet.iter_rows():
                data_row = []
                for cell in row:
                    data_row += [cell.value]
                data += [data_row]
                if only_header:
                    break
        elif file_ext == "ods":
            from .odsreader import ODSReader
            doc = ODSReader(self.import_file)
            table = doc.SHEETS.items()[0]

            # Remove blank columns that ods files seems to have
            blank_columns = []
            for i, header_cell in enumerate(table[1][0]):
                if self.is_empty(header_cell):
                    blank_columns += [i]
            # just an overly complicated way to remove these
            # indexes from a list
            for offset, blank_column in enumerate(blank_columns):
                for row in table[1]:
                    del row[blank_column - offset]

            if only_header:
                data += [table[1][0]]
            else:
                data += table[1]
        # Remove blank columns. We use the header row as a unique index. Can't handle blanks.
        columns_to_del = []
        for i, header_cell in enumerate(data[0]):
            if self.is_empty(header_cell):
                columns_to_del += [i]
        num_deleted = 0
        for column_to_del in columns_to_del:
            for row in data:
                del row[column_to_del - num_deleted]
            num_deleted += 1
        if only_header:
            return data[0]
        return data


class RelationalMatch(models.Model):
    """Store which unique field is being use to match.
    This can be used only to set a FK or one M2M relation
    on the import root model. It does not add them.
    With Multple rows set to the same field, you could set more
    than one per row.
    EX Lets say a student has an ID and username and both
    are marked as unique in Django orm. The user could reference
    that student by either."""
    import_log = models.ForeignKey(ImportLog)
    field_name = models.CharField(max_length=255) # Ex student_number_set
    related_field_name = models.CharField(max_length=255, blank=True) # Ex username


class ImportedObject(models.Model):
    import_log = models.ForeignKey(ImportLog)
    object_id = models.IntegerField()
    content_type = models.ForeignKey(ContentType)
    content_object = generic.GenericForeignKey('content_type', 'object_id')

